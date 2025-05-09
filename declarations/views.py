import datetime
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from exams.models import Grade
from users.models import StudentProfile
from .models import Course, ResitExamContent, ResitExamSchedule
from .forms import ExcelUploadForm
from django.contrib.auth.models import Group
import openpyxl
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from openpyxl.utils.datetime import from_excel



@login_required
def upload_resit_schedule(request):
    if not request.user.groups.filter(name='faculty').exists():
        return JsonResponse({'status': 'error', 'message': 'Unauthorized access.'}, status=403)

    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                excel_file = request.FILES['excel_file']
                if not excel_file.name.endswith(('.xlsx', '.xls')):
                    return JsonResponse({'status': 'error', 'message': 'Please upload a valid Excel file (.xlsx or .xls).'}, status=400)

                workbook = openpyxl.load_workbook(excel_file)
                worksheet = workbook.active
                headers = [cell.value.lower() if cell.value else '' for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
                expected_headers = ['course id', 'course name', 'place', 'date', 'time']

                if not all(header in headers for header in expected_headers):
                    return JsonResponse({
                        'status': 'error',
                        'message': 'Excel file must contain columns: Course ID, Course Name, Place, Date, Time.'
                    }, status=400)

                col_indices = {header: headers.index(header) for header in expected_headers}
                errors = []

                for row_idx, data_row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                    try:
                        course_id = data_row[col_indices['course id']].value
                        course_name = data_row[col_indices['course name']].value
                        place = data_row[col_indices['place']].value
                        date_raw = data_row[col_indices['date']].value
                        time_raw = data_row[col_indices['time']].value

                        if not all([course_id, course_name, place, date_raw, time_raw]):
                            errors.append(f"Row {row_idx}: All fields are required.")
                            continue

                        try:
                            course = Course.objects.get(code=str(course_id).strip())
                            if str(course_name).strip().lower() != course.name.lower():
                                errors.append(f"Row {row_idx}: Course name mismatch for ID {course_id}.")
                                continue
                        except Course.DoesNotExist:
                            errors.append(f"Row {row_idx}: Course with ID {course_id} not found.")
                            continue

                        # Robust date/time parsing
                        try:
                            if isinstance(date_raw, str):
                                date_parsed = datetime.strptime(date_raw.strip(), '%Y-%m-%d').date()
                            elif hasattr(date_raw, 'date'):
                                date_parsed = date_raw.date()
                            else:
                                date_parsed = date_raw

                            if isinstance(time_raw, str):
                                time_parsed = datetime.strptime(time_raw.strip(), '%H:%M').time()
                            elif isinstance(time_raw, (int, float)):
                                time_parsed = from_excel(time_raw).time()
                            elif hasattr(time_raw, 'time'):
                                time_parsed = time_raw.time()
                            else:
                                time_parsed = time_raw
                        except Exception as e:
                            errors.append(f"Row {row_idx}: Failed to parse date/time - {e}")
                            continue

                        ResitExamSchedule.objects.update_or_create(
                            course=course,
                            defaults={
                                'place': place.strip(),
                                'date': date_parsed,
                                'time': time_parsed
                            }
                        )
                    except Exception as e:
                        errors.append(f"Row {row_idx}: {str(e)}")

                if errors:
                    return JsonResponse({'status': 'error', 'message': '\n'.join(errors)}, status=400)

                return JsonResponse({'status': 'success', 'message': 'Resit exam schedule uploaded successfully.'})

            except Exception as e:
                return JsonResponse({'status': 'error', 'message': f'Error processing Excel file: {str(e)}'}, status=500)

        return JsonResponse({'status': 'error', 'message': 'Invalid form submission.'}, status=400)

    form = ExcelUploadForm()
    return render(request, 'facultysecexam.html', {'form': form})



@login_required
def get_resit_schedule(request):
    if not request.user.groups.filter(name='student').exists():
        return JsonResponse({'status': 'error', 'message': 'Unauthorized access.'}, status=403)

    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    schedule = {day: [] for day in days}  # ⬅ Make each day an empty list

    resits = ResitExamSchedule.objects.select_related('course')

    for resit in resits:
        try:
            day_name = resit.date.strftime('%A')
            if day_name in schedule:
                time_str = resit.time.strftime('%H:%M') if resit.time else 'Time not set'
                schedule[day_name].append({
                    'course': f"{resit.course.code} {resit.course.name}",
                    'date': f"{resit.date.strftime('%Y-%m-%d')} {time_str}",
                    'place': resit.place
                })
        except Exception:
            continue

    return JsonResponse({'status': 'success', 'schedule': schedule})


@require_POST
def upload_resit_details(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    excel_file = request.FILES.get('excel_file')

    if not excel_file:
        return JsonResponse({'status': 'error', 'message': 'No file uploaded'}, status=400)

    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) < 4:
                raise ValueError(f"Incomplete data in row: {row}")

            num_questions, exam_type, calculator_allowed, additional_notes = row

            if not isinstance(num_questions, (int, float)):
                raise ValueError(f"Invalid 'num_questions' value: {num_questions}")

            calc_allowed = str(calculator_allowed).strip().lower()
            if calc_allowed not in ['yes', 'no']:
                raise ValueError(f"Invalid 'calculator_allowed' value: {calculator_allowed}")
            calculator_bool = calc_allowed == 'yes'

            ResitExamContent.objects.update_or_create(
                course=course,
                defaults={
                    'num_questions': int(num_questions),
                    'exam_type': str(exam_type or '').strip(),
                    'calculator_allowed': calculator_bool,
                    'additional_notes': str(additional_notes or '').strip()
                }
            )

        return JsonResponse({'status': 'success'})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@login_required
def resitannouncement(request):
    if not request.user.groups.filter(name='student').exists():
        messages.error(request, "Only students can view resit exam announcements.")
        return redirect('student_page')

    try:
        student_profile = StudentProfile.objects.get(user=request.user)
    except StudentProfile.DoesNotExist:
        messages.error(request, "Student profile not found.")
        return redirect('student_page')

    declared_courses = Grade.objects.filter(
        student=student_profile,
        declared_resit=True
    ).values_list('course_id', flat=True)

    schedules = ResitExamSchedule.objects.filter(course__id__in=declared_courses).select_related('course')
    contents = ResitExamContent.objects.filter(course__id__in=declared_courses).select_related('course')

    resit_details = {}
    for sched in schedules:
        resit_details[sched.course.id] = {'schedule': sched}
    for content in contents:
        if content.course.id in resit_details:
            resit_details[content.course.id]['content'] = content
        else:
            resit_details[content.course.id] = {'content': content}

    return render(request, 'resitannouncement.html', {'resit_details': resit_details})





def facultysecexam(request):
    return render(request, 'facultysecexam.html')

@login_required
def studentshedule(request):
    if not request.user.groups.filter(name='student').exists():
        return JsonResponse({'status': 'error', 'message': 'Unauthorized access.'}, status=403)

    weekdays = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    schedule_by_day = {day: [] for day in weekdays}

    resits = ResitExamSchedule.objects.select_related('course')
    contents = ResitExamContent.objects.select_related('course')

    # Map content by course_id
    content_map = {c.course_id: c for c in contents}

    for resit in resits:
        day = resit.date.strftime('%A')
        if day in schedule_by_day:
            course_id = resit.course_id
            content = content_map.get(course_id)
            
            schedule_by_day[day].append({
                'course': f"{resit.course.code} {resit.course.name}",
                'date': resit.date.strftime('%Y-%m-%d'),
                'time': resit.time.strftime('%H:%M') if resit.time else '',
                'place': resit.place,
                'content': content  # this is safe to pass to the template
            })

    return render(request, 'studentschedule.html', {
        'weekdays': weekdays,
        'schedule_by_day': schedule_by_day
    })

